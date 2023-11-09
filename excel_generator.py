import pandas as pd
import os
import openpyxl
import json
#
# def generator():
#     # Создаем файл, задаем строки и столбцы
#     xlx = openpyxl.Workbook()
#     output_file = xlx.active
#     columns = ["Город/день", "2022-05-26", "2022-05-27", "2022-05-28",
#                "2022-05-29", "2022-05-30", "Среднее", "Результат"]
#     for col, column in enumerate(columns, start=1):
#         output_file.cell(row=1, column=col).value = column
#     xlx.save('merged_data.xlsx')
#     # Получаем список файлов с данными
#     files_path = ".\my_outputs"
#     files = [f for f in os.listdir(files_path) if f.endswith(".json")]
#
#     with pd.ExcelWriter('merged_data.xlsx', engine='openpyxl', mode='a') as writer:
#         for file in files:
#             json_path = os.path.join(files_path, file)
#             with open(json_path, "r") as f:
#                 json_data = pd.read_json(f)
#                 start_row = output_file.max_row + 1  # Получение следующей строки для записи данных
#                 json_data.to_excel(writer, index=False, header=False, startrow=start_row)
#                 writer.book.save('merged_data.xlsx')
#
#
# generator()


# Создаем новый xlsx файл
wb = openpyxl.Workbook()
sheet = wb.active

# Открываем JSON-файл и загружаем данные
with open('.\my_outputs\BERLIN_output.json', 'r') as f:
    data = json.load(f)
    days = data['days']

# Записываем ключи-значения в отдельный файл Excel
row = 1
for day in days:
    col = 1
    for key, value in day.items():
        sheet.cell(row=row, column=col).value = key
        sheet.cell(row=row+1, column=col).value = value
        col += 1
        row += 2

# Сохраняем файл Excel
wb.save('data.xlsx')
